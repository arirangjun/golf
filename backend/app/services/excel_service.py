from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services.booking_rules import format_hour
from app.services.formatting import format_member_display, format_phone, format_unit
from app.services.member_service import ImportRow


def build_member_template_buffer() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "회원등록"
    ws.append(["동", "호수", "이름", "휴대폰", "비밀번호"])
    ws.append(["101", "1001", "홍길동", "01012345678", "1"])
    ws.append(["101", "1001", "김영희", "01098765432", "2"])
    ws.append(["102", "2001", "김철수", "01055556666", "1"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _normalize_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def _cell_value(row: list, index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def parse_member_excel(content: bytes) -> list[ImportRow]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = [_normalize_header(cell) for cell in rows[0]]
    dong_idx = next((i for i, h in enumerate(header_row) if h in ("동", "dong")), -1)
    ho_idx = next((i for i, h in enumerate(header_row) if h in ("호수", "ho", "호")), -1)
    name_idx = next((i for i, h in enumerate(header_row) if h in ("이름", "name", "성명")), -1)
    phone_idx = next(
        (i for i, h in enumerate(header_row) if h in ("휴대폰", "휴대폰번호", "phone", "연락처", "전화번호")),
        -1,
    )
    password_idx = next(
        (i for i, h in enumerate(header_row) if h in ("비밀번호", "password", "pw")), -1
    )

    has_header = dong_idx >= 0 and ho_idx >= 0 and name_idx >= 0
    if has_header:
        indices = {"dong": dong_idx, "ho": ho_idx, "name": name_idx, "phone": phone_idx, "password": password_idx}
        start = 1
    else:
        indices = {"dong": 0, "ho": 1, "name": 2, "phone": 3, "password": 4}
        start = 0

    result: list[ImportRow] = []
    for i in range(start, len(rows)):
        row = list(rows[i])
        dong = _cell_value(row, indices["dong"])
        ho = _cell_value(row, indices["ho"])
        name = _cell_value(row, indices["name"])
        phone = _cell_value(row, indices["phone"]) if indices["phone"] >= 0 else ""
        password = _cell_value(row, indices["password"]) if indices["password"] >= 0 else ""

        if not dong and not ho and not name and not phone and not password:
            continue

        result.append(
            ImportRow(
                row=i + 1,
                dong=dong,
                ho=ho,
                name=name,
                phone=phone,
                password=password or None,
            )
        )

    return result


def build_stats_export_buffer(data: dict) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "요약"
    summary.append(["항목", "값"])
    summary.append(["조회 시작", data["from"]])
    summary.append(["조회 종료", data["to"]])
    summary.append(["총 예약 건수", data["total"]])

    member_stats = data.get("memberStats")
    if member_stats:
        summary.append(["월별 회원 집계 월", member_stats["month"]])
        summary.append(["해당 월 예약 회원 수", member_stats["uniqueMembers"]])
        summary.append(["해당 월 총 예약", member_stats["totalReservations"]])

    def add_sheet(title: str, headers: list, rows: list):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    add_sheet("일별", ["날짜", "예약 건수"], [[d["label"], d["count"]] for d in data["daily"]])
    add_sheet("주별", ["주 시작일", "예약 건수"], [[d["label"], d["count"]] for d in data["weekly"]])
    add_sheet("월별", ["월", "예약 건수"], [[d["label"], d["count"]] for d in data["monthly"]])
    add_sheet(
        "타임별",
        ["시간", "예약 건수", "이용률(%)"],
        [[format_hour(h["hour"]), h["count"], h["rate"]] for h in data["hourlyUtilization"]],
    )

    if member_stats:
        add_sheet(
            "월별회원집계",
            ["순위", "동", "이름(마스킹)", "휴대폰", "예약 횟수"],
            [
                [idx + 1, m["dong"], m["displayName"], m["phone"], m["count"]]
                for idx, m in enumerate(member_stats["members"])
            ],
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
